import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

products_per_supplier = {}
totalvalue = {}
inventory_lessthan_10 = {}


for product_row in range(2, product_list.max_row + 1):  #starting from 2nd row, product_list.max_row isnt last row, so we should add +1 for including last row"
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

 #calculation product per supplier
    if supplier_name in products_per_supplier:
        products_per_supplier[supplier_name] = products_per_supplier[supplier_name] + 1
    else:
        products_per_supplier[supplier_name] = 1

 #calculation total value of inventory per supplier
    if supplier_name in totalvalue:
        totalvalue[supplier_name] = totalvalue.get(supplier_name) + inventory * price
    else:
        totalvalue[supplier_name] = inventory * price

 #counting inventories less than 10
    if inventory < 10:
        inventory_lessthan_10[product] = inventory

 #adding a value for total inventory price
    inventory_price.value = inventory * price

print(products_per_supplier)
print(totalvalue)
print(inventory_lessthan_10)

inv_file.save("inventory_with_total_value.xlsx")
